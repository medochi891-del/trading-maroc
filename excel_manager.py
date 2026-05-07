"""
excel_manager.py
Gestion du fichier Excel - Mohamed Fadel Babiya
Design professionnel sans emojis
"""

import os
import shutil
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_FILE  = "Mohamed_Fadel_Babiya_Trading_Simulation.xlsx"
ARCHIVE_DIR = "archive"

# Couleurs
C_BLEU_FONCE = "1F3864"
C_BLEU_MOYEN = "2E75B6"
C_BLEU_CLAIR = "D6E4F0"
C_VERT       = "C6EFCE"
C_ROUGE      = "FFC7CE"
C_JAUNE      = "FFEB9C"
C_BLEU_ATT   = "DDEBF7"
C_BLANC      = "FFFFFF"
C_GRIS       = "F2F2F2"
C_VERT_TXT   = "1F6B3A"
C_ROUGE_TXT  = "9C0006"

def _bord():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _entete(ws, row, col, value, bg=C_BLEU_FONCE, size=11,
            wrap=False, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=C_BLANC, size=size, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = _bord()
    return c

def _cellule(ws, row, col, value, bg=C_BLANC, bold=False,
             align="center", color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, size=10, name="Calibri", color=color)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border    = _bord()
    return c


# ═══════════════════════════════════════════════════
# CREATION DU FICHIER EXCEL
# ═══════════════════════════════════════════════════
def create_excel(config: dict):
    wb = Workbook()
    _page_donnee(wb, config)
    _page_suivi(wb)
    _page_presentation(wb)
    wb.save(EXCEL_FILE)
    print(f"Fichier Excel cree : {EXCEL_FILE}")


def _page_donnee(wb, config):
    ws = wb.active
    ws.title = "Donnee"
    ws.sheet_view.showGridLines = False

    # Titre principal
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value     = "SIMULATION DE TRADING - BOURSE DE CASABLANCA"
    c.font      = Font(bold=True, size=14, color=C_BLANC, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_BLEU_FONCE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Nom etudiant
    ws.merge_cells("A2:C2")
    c = ws["A2"]
    c.value     = "Mohamed Fadel Babiya"
    c.font      = Font(bold=True, size=12, color=C_BLEU_FONCE, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_BLEU_CLAIR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 26

    # En-tetes colonnes
    for j, label in enumerate(["Etape", "Description", "Detail"], 1):
        _entete(ws, 3, j, label, bg=C_BLEU_MOYEN, size=10)
    ws.row_dimensions[3].height = 22

    # Contenu
    lignes = [
        ("ETAPE 1",
         "Choix de l'entreprise",
         f"{config['stock_name']} ({config['ticker']}) - {config['sector']}. "
         f"Justification : {config['justification']}"),

        ("ETAPE 2",
         "Strategie de trading",
         "Strategie de suivi de tendance (Moyennes Mobiles, MACD) et "
         "strategie contre-tendance (RSI, Bandes de Bollinger, Chandeliers japonais). "
         "La strategie peut etre modifiee chaque jour selon l'evolution du marche."),

        ("ETAPE 3",
         "Analyse de la situation",
         "Analyse technique complete basee sur 7 indicateurs du cours : "
         "RSI, Moyennes Mobiles (MM5/MM10), MACD, Bandes de Bollinger, "
         "Chandeliers japonais, Supports et Resistances, Figure ETE/ETEI. "
         "Une decision est prise si minimum 3 indicateurs confirment le signal."),

        ("ETAPE 4",
         "Decision d'investissement",
         f"Capital disponible : {config['capital']:,} MAD. "
         f"Montant investi : 20% du capital actuel. "
         f"Objectif de vente : Take-Profit a +4%. "
         f"Maximum 2 operations par jour."),

        ("ETAPE 5",
         "Stop-Loss et Take-Profit",
         "Stop-Loss : -2% du prix d'achat (protection du capital). "
         "Take-Profit : +4% du prix d'achat (securisation des gains). "
         "Ratio gain/risque : 1:2. "
         "Sortie anticipee possible si 3 signaux de danger sont detectes."),

        ("CAPITAL",
         "Capital initial",
         f"{config['capital']:,} MAD"),

        ("PERIODE",
         "Duree de simulation",
         "10 jours ouvrables (2 semaines calendaires - weekends exclus)"),

        ("SOURCE",
         "Source des cours",
         "Yahoo Finance (donnees OHLCV historiques) + "
         "casablancabourse.com (prix en temps reel)"),
    ]

    for i, (etape, desc, detail) in enumerate(lignes, start=4):
        bg = C_BLEU_CLAIR if i % 2 == 0 else C_BLANC

        c1 = ws.cell(row=i, column=1, value=etape)
        c1.font      = Font(bold=True, size=10, color=C_BLANC, name="Calibri")
        c1.fill      = PatternFill("solid", fgColor=C_BLEU_MOYEN)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border    = _bord()

        c2 = ws.cell(row=i, column=2, value=desc)
        c2.font      = Font(bold=True, size=10, name="Calibri")
        c2.fill      = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c2.border    = _bord()

        c3 = ws.cell(row=i, column=3, value=detail)
        c3.font      = Font(size=10, name="Calibri")
        c3.fill      = PatternFill("solid", fgColor=C_BLANC)
        c3.alignment = Alignment(horizontal="left", vertical="center",
                                  wrap_text=True, indent=1)
        c3.border    = _bord()
        ws.row_dimensions[i].height = 46

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 78


def _page_suivi(wb):
    ws = wb.create_sheet("Suivi Portefeuille")
    ws.sheet_view.showGridLines = False

    # Titre
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = "SUIVI DU PORTEFEUILLE - 10 JOURS OUVRABLES"
    c.font      = Font(bold=True, size=13, color=C_BLANC, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_BLEU_FONCE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # En-tetes
    headers = [
        "Date", "Actif",
        "Prix d'achat\n(MAD)", "Prix actuel\n(MAD)",
        "Quantite", "Valeur\n(MAD)", "Gain / Perte\n(MAD)",
        "Strategie", "Explication des decisions"
    ]
    widths = [13, 18, 15, 14, 10, 14, 16, 25, 75]

    for j, (h, w) in enumerate(zip(headers, widths), 1):
        _entete(ws, 2, j, h, bg=C_BLEU_MOYEN, wrap=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[2].height = 36

    # Legende sans emojis
    ws.merge_cells("A33:I33")
    leg = ws["A33"]
    leg.value = (
        "Couleurs : Vert = ACHETER | Rouge = VENDRE | "
        "Jaune = CONSERVER | Bleu clair = ATTENDRE"
    )
    leg.font      = Font(size=9, italic=True, color="595959", name="Calibri")
    leg.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[33].height = 18


def _page_presentation(wb):
    ws = wb.create_sheet("Presentation d'analyse et Rtat")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value     = "PRESENTATION FINALE - RESULTATS DE LA SIMULATION"
    c.font      = Font(bold=True, size=13, color=C_BLANC, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_BLEU_FONCE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:C2")
    c = ws["A2"]
    c.value     = "A completer automatiquement au terme des 10 jours de simulation"
    c.font      = Font(italic=True, size=10, color="595959", name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_GRIS)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    labels = [
        "Strategie utilisee",
        "Gestion du risque",
        "Capital initial (MAD)",
        "Montant total investi (MAD)",
        "Nombre total d'operations",
        "Operations gagnantes",
        "Operations perdantes",
        "Gain / Perte total(e) (MAD)",
        "Capital final (MAD)",
        "Performance (%)",
    ]

    for i, label in enumerate(labels, start=3):
        bg = C_BLEU_CLAIR if i % 2 == 0 else C_BLANC

        c2 = ws.cell(row=i, column=2, value=label)
        c2.font      = Font(bold=True, size=11, name="Calibri")
        c2.fill      = PatternFill("solid", fgColor=bg)
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c2.border    = _bord()

        c3 = ws.cell(row=i, column=3, value="-")
        c3.font      = Font(size=11, name="Calibri")
        c3.fill      = PatternFill("solid", fgColor=C_BLANC)
        c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c3.border    = _bord()
        ws.row_dimensions[i].height = 26

    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 50


# ═══════════════════════════════════════════════════
# AJOUT D'UNE LIGNE DANS SUIVI PORTEFEUILLE
# ═══════════════════════════════════════════════════
def add_trade_row(jour: int, trade_num: int, date_str: str,
                  action: str, prix_achat, prix_actuel,
                  quantite, valeur, gain_perte,
                  strategie: str, explication: str):

    wb  = load_workbook(EXCEL_FILE)
    ws  = wb["Suivi Portefeuille"]

    # Trouver la prochaine ligne vide (commence a la ligne 3)
    next_row = 3
    for r in range(3, 200):   # 200 lignes max - largement suffisant
        if ws.cell(row=r, column=1).value is None:
            next_row = r
            break

    # Couleur selon action
    palette = {
        "ACHETER":   C_VERT,
        "VENDRE":    C_ROUGE,
        "CONSERVER": C_JAUNE,
        "ATTENDRE":  C_BLEU_ATT,
    }
    bg = palette.get(action, C_BLANC)

    # Label du jour
    label = f"J{jour}" if trade_num == 1 else f"J{jour} - Op.{trade_num}"

    # Couleur texte Gain/Perte
    gp_color = C_BLANC
    if gain_perte is not None and isinstance(gain_perte, (int, float)):
        gp_color = C_VERT_TXT if gain_perte >= 0 else C_ROUGE_TXT
        gain_perte = f"{gain_perte:+.2f}"

    valeurs = [
        label,
        "Itissalat Al-Maghrib (IAM)",
        f"{prix_achat:.2f}"  if isinstance(prix_achat,  float) else "-",
        f"{prix_actuel:.2f}" if isinstance(prix_actuel, float) else "-",
        str(quantite)        if quantite  is not None else "-",
        f"{valeur:.2f}"      if isinstance(valeur,      float) else "-",
        gain_perte           if gain_perte is not None  else "-",
        strategie,
        explication,
    ]

    for col, val in enumerate(valeurs, start=1):
        c = ws.cell(row=next_row, column=col, value=val)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(
            size  = 10,
            name  = "Calibri",
            color = gp_color if col == 7 else "000000"
        )
        c.alignment = Alignment(
            wrap_text  = True,
            vertical   = "top",
            horizontal = "left" if col == 9 else "center"
        )
        c.border = _bord()

    ws.row_dimensions[next_row].height = 58
    wb.save(EXCEL_FILE)
    print(f"Excel mis a jour - J{jour} Op{trade_num} : {action}")


# ═══════════════════════════════════════════════════
# PAGE PRESENTATION - RESULTATS FINAUX
# ═══════════════════════════════════════════════════
def fill_presentation(config: dict, total_gain: float,
                       nb_ops: int, nb_win: int, nb_loss: int):

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Presentation d'analyse et Rtat"]

    # Capital final = capital courant (deja mis a jour avec les gains/pertes)
    capital_final = config["capital"]
    capital_init  = 100000
    performance   = ((capital_final - capital_init) / capital_init) * 100

    resultats = [
        "Strategie de suivi de tendance + Strategie contre-tendance "
        "(changement selon les conditions du marche)",

        "Stop-Loss : -2% du prix d'achat | Take-Profit : +4% | "
        "Ratio gain/risque : 1:2 | Maximum 2 operations par jour",

        f"{capital_init:,} MAD",
        f"{config.get('invest_amount', 20000):,} MAD (20% du capital actuel)",
        str(nb_ops),
        str(nb_win),
        str(nb_loss),
        f"{total_gain:+,.2f} MAD",
        f"{capital_final:,.2f} MAD",
        f"{performance:+.2f}%",
    ]

    for i, val in enumerate(resultats, start=3):
        is_positive = isinstance(val, str) and "+" in val
        is_negative = isinstance(val, str) and val.startswith("-")
        color = C_VERT_TXT if is_positive else C_ROUGE_TXT if is_negative else "000000"

        c = ws.cell(row=i, column=3, value=val)
        c.font      = Font(size=11, name="Calibri", color=color)
        c.fill      = PatternFill("solid", fgColor=C_BLANC)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = _bord()

    wb.save(EXCEL_FILE)
    print("Page Presentation completee")


# ═══════════════════════════════════════════════════
# ARCHIVE JOURNALIERE
# ═══════════════════════════════════════════════════
def save_archive(jour: int):
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    today   = date.today().strftime("%d-%m-%Y")
    archive = os.path.join(
        ARCHIVE_DIR,
        f"Mohamed_Fadel_Babiya_J{jour}_{today}.xlsx"
    )
    shutil.copy2(EXCEL_FILE, archive)
    print(f"Archive sauvegardee : {archive}")
